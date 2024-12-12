
from openpyxl import workbook, load_workbook


def read_data_from_excel(file_name, sheet):
    data_list=[]
    
    wb=load_workbook(filename=file_name)
    sh=wb.active  #sh=wb["sheet_name"] ....if we have multiple sheets on a workbook

    row_ct=sh.max_row
    col_ct=sh.max_column

    for i in range(2, row_ct+1):
        row=[]
        for j in range(5, 7):
            row.append(sh.cell(row=i, column=j).value)
        data_list.append(row)
        
    return data_list

